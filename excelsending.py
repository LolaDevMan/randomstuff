# import openpyxl and tkinter modules 
from openpyxl import *
from tkinter import *
import matplotlib.pyplot as plt
import numpy as np

# opening the existing excel file 
wb = load_workbook('C:\\Users\\alwin\\Desktop\\form.xlsx') 

# create the sheet object 
sheet = wb.active 

def plotage():
    a=b=c=d=e=f=g=h=k=j=0
    for i in range(2,20):
        if sheet.cell(row=i, column=2).value is None:
            break
        elif int(sheet.cell(row=i, column=2).value) <= 10:
            a+=1
        elif int(sheet.cell(row=i, column=2).value) <= 20:
            b+=1
        elif int(sheet.cell(row=i, column=2).value) <= 30:
            c+=1
        elif int(sheet.cell(row=i, column=2).value) <= 40:
            d+=1
        elif int(sheet.cell(row=i, column=2).value) <= 50:
            e+=1
        elif int(sheet.cell(row=i, column=2).value) <= 60:
            f+=1
        elif int(sheet.cell(row=i, column=2).value)<= 70:
            g+=1
        elif int(sheet.cell(row=i, column=2).value) <= 80:
            h+=1
        elif int(sheet.cell(row=i, column=2).value) <= 90:
            k+=1
        else:
            j+=1
    x = np.array(["0-10","11-20","21-30","31-40","41-50","51-60","61-70","71-80","81-90","90+"])
    y = np.array([a,b,c,d,e,f,g,h,k,j])
    plt.bar(x,y)
    plt.show()
    
def plotbld():
    a=b=o=ab=aa=bb=oo=abab=0
    for i in range(2,20):
        if sheet.cell(row=i, column=3).value == "a+ve":
            a+=1
        elif sheet.cell(row=i, column=3).value == "b+ve":
            b+=1
        elif sheet.cell(row=i, column=3).value == "o+ve":
            o+=1
        elif sheet.cell(row=i, column=3).value == "ab+ve":
            ab+=1
        elif sheet.cell(row=i, column=3).value == "a-ve":
            aa+=1
        elif sheet.cell(row=i, column=3).value == "b-ve":
            bb+=1
        elif sheet.cell(row=i, column=3).value == "o-ve":
            oo+=1
        elif sheet.cell(row=i, column=3).value == "ab-ve":
            abab+=1
    x = np.array(["A+ve", "B+ve", "O+ve", "AB+ve", "A-ve", "B-ve", "O-ve", "AB-ve"])
    y = np.array([a,b,o,ab,aa,bb,oo,abab])
    plt.bar(x,y)
    plt.show()

def plotdis():
    dis = []
    for i in range(2,20):
        if sheet.cell(row=i, column=4).value is None:
            break
        else:
            dis.append(sheet.cell(row=i, column=4).value)
    dicti = {i:dis.count(i) for i in dis}
    xcoor = []
    ycoor = []
    items = dicti.items() 
    for item in items: 
        xcoor.append(item[0]), ycoor.append(item[1]) 
    x = np.array(xcoor)
    y = np.array(ycoor)
    plt.bar(x,y)
    plt.show()

def excel(): 
    
    # resize the width of columns in 
    # excel spreadsheet 
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    # write given data to an excel spreadsheet 
    # at particular location 
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Age"
    sheet.cell(row=1, column=3).value = "Blood Group"
    sheet.cell(row=1, column=4).value = "Disease Contracted"
    sheet.cell(row=1, column=5).value = "Phone Number"
    sheet.cell(row=1, column=6).value = "Guardian"
    sheet.cell(row=1, column=7).value = "Allergies(if any)"


# Function to set focus (cursor) 
def focus1(event): 
    # set focus on the age_field box 
    age_field.focus_set() 


# Function to set focus 
def focus2(event): 
    # set focus on the bld_field box 
    bld_field.focus_set() 


# Function to set focus 
def focus3(event): 
    # set focus on the dis_field box 
    dis_field.focus_set() 


# Function to set focus 
def focus4(event): 
    # set focus on the phn_field box 
    phn_field.focus_set() 


# Function to set focus 
def focus5(event): 
    # set focus on the guard_field box 
    guard_field.focus_set() 


# Function to set focus 
def focus6(event): 
    # set focus on the allergy_field box 
    allergy_field.focus_set() 


# Function for clearing the 
# contents of text entry boxes 
def clear(): 
    
    # clear the content of text entry box 
    name_field.delete(0, END) 
    age_field.delete(0, END) 
    bld_field.delete(0, END) 
    dis_field.delete(0, END) 
    phn_field.delete(0, END) 
    guard_field.delete(0, END) 
    allergy_field.delete(0, END) 


# Function to take data from GUI 
# window and write to an excel file 
def insert(): 
    
    # if user not fill any entry 
    # then print "empty input" 
    if (name_field.get() == "" and
        age_field.get() == "" and
        bld_field.get() == "" and
        dis_field.get() == "" and
        phn_field.get() == "" and
        guard_field.get() == "" and
        allergy_field.get() == ""): 
            
        print("empty input") 

    else: 

        # assigning the max row and max column 
        # value upto which data is written 
        # in an excel sheet to the variable 
        current_row = sheet.max_row 
        current_column = sheet.max_column 

        # get method returns current text 
        # as string which we write into 
        # excel spreadsheet at particular location 
        sheet.cell(row=current_row + 1, column=1).value = name_field.get() 
        sheet.cell(row=current_row + 1, column=2).value = age_field.get() 
        sheet.cell(row=current_row + 1, column=3).value = bld_field.get() 
        sheet.cell(row=current_row + 1, column=4).value = dis_field.get() 
        sheet.cell(row=current_row + 1, column=5).value = phn_field.get() 
        sheet.cell(row=current_row + 1, column=6).value = guard_field.get() 
        sheet.cell(row=current_row + 1, column=7).value = allergy_field.get() 

        # save the file 
        wb.save('C:\\Users\\alwin\\Desktop\\form.xlsx') 

        # set focus on the name_field box 
        name_field.focus_set() 

        # call the clear() function 
        clear() 

def openNewWindow(): 
      
    # Toplevel object which will  
    # be treated as a new window 
    newWindow = Toplevel(root)

    newWindow.configure(background="wheat1") 

    
    # sets the title of the 
    # Toplevel widget 
    newWindow.title("Switcher") 
  
    # sets the geometry of toplevel 
    newWindow.geometry("350x200") 
    lnew = Label(newWindow, text="          Click on what you want the graph to be plotted on", bg="wheat1")
    lnew.grid(row=0, column=1) 
    sp1 = Label(newWindow, text=" ", bg="wheat1")
    sp1.grid(row=1, column=1) 
    agebtn = Button(newWindow, text="  Age  ", fg="Black", bg="burlywood1",command=plotage)
    agebtn.grid(row=2, column=1) 
    sp2 = Label(newWindow, text=" ", bg="wheat1")
    sp2.grid(row=3, column=1) 
    bldbtn = Button(newWindow, text="Blood Group", fg="Black", bg="burlywood1",command=plotbld)
    bldbtn.grid(row=4, column=1) 
    sp3 = Label(newWindow, text=" ", bg="wheat1")
    sp3.grid(row=5, column=1) 
    disbtn = Button(newWindow, text=" Disease ", fg="Black", bg="burlywood1",command=plotdis)
    disbtn.grid(row=6, column=1) 

if __name__ == "__main__": 
    
    root = Tk() 

    root.configure(background="peach puff") 

    root.title("registration form") 

    root.geometry("500x300") 

    excel() 

    heading = Label(root, text="Patient Details", bg="peach puff") 
    
    name = Label(root, text="Name", bg="peach puff")

    age = Label(root, text="Age", bg="peach puff") 

    bld = Label(root, text="Blood Group", bg="peach puff") 

    dis = Label(root, text="Disease Contracted", bg="peach puff") 

    phn = Label(root, text="Phone Number", bg="peach puff") 

    guard = Label(root, text="Guardian", bg="peach puff") 

    allergy = Label(root, text="Allergies(if any)", bg="peach puff") 

    heading.grid(row=0, column=1) 
    name.grid(row=1, column=0) 
    age.grid(row=2, column=0) 
    bld.grid(row=3, column=0) 
    dis.grid(row=4, column=0) 
    phn.grid(row=5, column=0) 
    guard.grid(row=6, column=0) 
    allergy.grid(row=7, column=0) 

    name_field = Entry(root) 
    age_field = Entry(root) 
    bld_field = Entry(root) 
    dis_field = Entry(root) 
    phn_field = Entry(root) 
    guard_field = Entry(root) 
    allergy_field = Entry(root) 

    # bind method of widget is used for 
    # the binding the function with the events 

    name_field.bind("<Return>", focus1) 

    age_field.bind("<Return>", focus2) 

    bld_field.bind("<Return>", focus3) 

    dis_field.bind("<Return>", focus4) 

    phn_field.bind("<Return>", focus5) 

    guard_field.bind("<Return>", focus6) 
 
    name_field.grid(row=1, column=1, ipadx="100") 
    age_field.grid(row=2, column=1, ipadx="100") 
    bld_field.grid(row=3, column=1, ipadx="100") 
    dis_field.grid(row=4, column=1, ipadx="100") 
    phn_field.grid(row=5, column=1, ipadx="100") 
    guard_field.grid(row=6, column=1, ipadx="100") 
    allergy_field.grid(row=7, column=1, ipadx="100") 

    excel() 
    l = Label(root, text=" ", bg="peach puff") 
    l.grid(row=8,column=1) 

    submit = Button(root, text="  Submit  ", fg="Black", 
                            bg="pink", command=insert) 
    submit.grid(row=9, column=1)

    btn = Button(root,  
             text ="  Finish  ", 
             fg="Black", 
             bg="pink", 
             command = openNewWindow) 
    
    btn.grid(row=11, column=1) 

    root.mainloop() 
 
